import openpyxl
import os
import glob


files = list()
def get_files():
    path = os.getcwd()  # 获取当前路径
    dir_file = glob.glob(path+"\\*.xlsx")
    for filenames in dir_file:
        file = filenames.split('\\')[-1]
        files.append(file)
        print(file)


files_data = list()
def get_a_file_data(file):
    wb = openpyxl.load_workbook(file)  # open excel
    sheet_names = wb.sheetnames  # 表格所有列表名称
    print(sheet_names)  # 打印sheet名称
    all_sheet = list()
    for sheet in sheet_names:
        all_sheet.append(sheet)
        print(sheet)
    ws = wb[all_sheet[0]] # 获取第一张表
    # print(ws)  # 打印选中的表格对象
    print('行总数：' + str(ws.max_row))  # 行总数
    print(ws.max_column)  # 列总数
    print('列总数：' + str(ws.max_column))
    for j in ws.rows:  # we.rows 获取每一行数据
        row_data = dict()
        a = 1
        for n in j:
            print(n.value, end="\t")  # n.value 获取单元格的值
            row_data[a] = n.value
            a = a + 1
        files_data.append(row_data)
        print()


def analysis():
    # 获取文件名 #xx.py xx.xlsx
    # file = sys.argv[1]
    file = input('File name: ')
    print(file)
    get_a_file_data(file)


if __name__ == "__main__":
    analysis()